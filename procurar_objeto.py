"""
procurar_objeto.py — Módulo de Processamento de Exploração de Objetos

Filtra, organiza e formata dados de exploração de objetos gerados pelo
software Topscan. Organiza pares de objetos lado a lado, calculando
métricas como soma de duração total e índice de discriminação (DI).
"""

from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ─── Constantes ───────────────────────────────────────────────
TOPSCAN_HEADER_ROW = 6
IDENTITY_COLUMNS = ['DAY', 'ANIMAL', 'OBJECTS']

COLUMN_RENAMES = {
    'OBJECTS': 'Objetos',
    'DAY': 'Dia',
    'ANIMAL': 'Animal',
    'Total Bouts': 'Bouts',
    'Total Duration(Second)': 'Exploração',
    'Blank Column 1': 'Total',
    'Blank Column 2': 'DI',
    'Latency(Second)': 'Latência',
    'Ending time(Second) of First Bout': 'FIM',
}

# Mapeamento de colunas Excel para merge de cabeçalhos (OBJ1 | OBJ2)
MERGE_COLUMN_PAIRS = [('D', 'E'), ('F', 'G'), ('J', 'K'), ('L', 'M')]
SEPARATOR_COLUMNS = ['H', 'I']
MERGE_COLUMN_WIDTH = 12
SEPARATOR_COLUMN_WIDTH = 16

# Estilos do Excel
CENTERED_ALIGNMENT = Alignment(horizontal='center', vertical='center')
HEADER_FONT = Font(bold=True)
HEADER_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def procurar(
    primeiro_objeto: str,
    segundo_objeto: str,
    primeiro_obj: str,
    segundo_obj: str,
    caminho_arquivo1: str,
    global_workbook: Workbook,
    colunas_desejadas: List[str],
) -> Tuple[str, List[str]]:
    """
    Processa dados de exploração de objetos do Topscan.

    Filtra os eventos de sniffing para cada par de objetos, organiza lado a lado,
    calcula Total e DI (índice de discriminação), e salva em uma nova aba do workbook.

    Args:
        primeiro_objeto: Identificador do primeiro par de objeto (ex: 'A').
        segundo_objeto:  Identificador do segundo par de objeto (ex: 'B').
        primeiro_obj:    Número do primeiro OBJ (ex: '1').
        segundo_obj:     Número do segundo OBJ (ex: '2').
        caminho_arquivo1: Caminho completo do arquivo Excel do Topscan.
        global_workbook:  Workbook do openpyxl para escrita.
        colunas_desejadas: Lista de colunas a extrair da planilha.

    Returns:
        Tupla com (objeto_desejado, lista_de_eventos).
    """
    # ─── Carregamento e Normalização ──────────────────────────
    df = pd.read_excel(caminho_arquivo1, header=TOPSCAN_HEADER_ROW)

    # Garante que colunas de string não contenham floats NaN antes de qualquer filtragem
    df['OBJECTS'] = df['OBJECTS'].fillna('')
    df['Events'] = df['Events'].fillna('')

    df['DAY'] = df['DAY'].astype(str)
    df['DAY'] = df['DAY'].apply(lambda x: str(x)[:-2] if str(x).endswith('.0') else str(x))
    df['DAY'] = df['DAY'].map(lambda x: int(x) if x.isdigit() else x)

    objeto_desejado = f'{primeiro_objeto}{segundo_objeto}'
    eventos_desejados = [f'OBJ{primeiro_obj}', f'OBJ{segundo_obj}']

    ws = global_workbook.create_sheet(title=f'{primeiro_objeto}_{segundo_objeto}')

    # ─── Processamento dos Eventos ────────────────────────────
    dfs_finais = []
    df_primeiro_objeto = None

    if df_primeiro_objeto is None:
        df_primeiro_objeto = pd.DataFrame(columns=colunas_desejadas)

    for evento_desejado in eventos_desejados:
        evento = f'Mouse 1 sniffing On {evento_desejado}'
        dados_do_evento = df[(df['Events'] == evento) & (df['OBJECTS'].str.contains(objeto_desejado, na=False))]

        colunas_desejadas = ['DAY', 'ANIMAL', 'OBJECTS', 'Total Bouts', 'Total Duration(Second)', 'Latency(Second)', 'Ending time(Second) of First Bout']

        dados_filtrados = []
        ultimo_dia = None

        # Filtra e organiza os dados com separadores entre dias
        for _, row in dados_do_evento.iterrows():
            if ultimo_dia is not None and row['DAY'] != ultimo_dia:
                dados_filtrados.append([None] * len(colunas_desejadas))

            linha = []
            for coluna in colunas_desejadas:
                linha.append(row[coluna])
                # Coluna auxiliar em branco para dados do segundo objeto
                if coluna not in IDENTITY_COLUMNS:
                    linha.append(None)

            dados_filtrados.append(linha)
            ultimo_dia = row['DAY']

        # Monta colunas expandidas (col + col_sufixo para cada coluna de dados)
        colunas_finais = []
        for col in colunas_desejadas:
            colunas_finais.append(col)
            if col not in IDENTITY_COLUMNS:
                colunas_finais.append(f"{col}_{segundo_objeto}")
        df_final = pd.DataFrame(dados_filtrados, columns=colunas_finais)

        # ─── Primeira Iteração: armazena dados do OBJ1 ────────
        if evento_desejado == f'OBJ{primeiro_obj}':
            df_primeiro_objeto = df_final

        # ─── Segunda Iteração: mescla dados do OBJ2 no OBJ1 ──
        else:
            for _, row in df_final.iterrows():
                idx = df_primeiro_objeto[
                    (df_primeiro_objeto['DAY'] == row['DAY']) &
                    (df_primeiro_objeto['ANIMAL'] == row['ANIMAL'])
                ].index

                for i in idx:
                    # Preenche colunas auxiliares com dados do segundo objeto
                    for col in colunas_desejadas[3:]:
                        df_primeiro_objeto.at[i, f"{col}_{segundo_objeto}"] = row[col]

                    idx_total_duration = df_primeiro_objeto.columns.get_loc(f"Total Duration(Second)_{segundo_objeto}")

                    # Insere colunas de cálculo: Total e DI
                    if 'Blank Column 1' not in df_primeiro_objeto.columns:
                        df_primeiro_objeto.insert(idx_total_duration + 1, 'Blank Column 1', None)
                    if 'Blank Column 2' not in df_primeiro_objeto.columns:
                        df_primeiro_objeto.insert(idx_total_duration + 2, 'Blank Column 2', None)

                    # Cálculo: DI = (dur2 - dur1) / (dur1 + dur2)
                    subtrair = df_primeiro_objeto.at[i, f"Total Duration(Second)_{segundo_objeto}"] - df_primeiro_objeto.at[i, "Total Duration(Second)"]
                    soma = df_primeiro_objeto.at[i, "Total Duration(Second)"] + df_primeiro_objeto.at[i, f"Total Duration(Second)_{segundo_objeto}"]
                    if pd.notna(subtrair) and pd.notna(soma) and soma != 0:
                        df_primeiro_objeto.at[i, 'Blank Column 1'] = soma
                        df_primeiro_objeto.at[i, 'Blank Column 2'] = round(subtrair / soma, 2)
                    else:
                        df_primeiro_objeto.at[i, 'Blank Column 1'] = float('nan')
                        df_primeiro_objeto.at[i, 'Blank Column 2'] = float('nan')

    # ─── Concatenação e Coluna de Droga ───────────────────────
    dfs_finais.append(df_primeiro_objeto)
    df_final_concatenado = pd.concat(dfs_finais, axis=1)

    df['Events'] = df['Events'].fillna('')

    valores_droga = []
    for _, row in df_final_concatenado.iterrows():
        evento_primeiro_objeto = f'Mouse 1 sniffing On OBJ{primeiro_obj}'
        linha_correspondente = df[
            (df['Events'] == evento_primeiro_objeto) &
            (df['DAY'] == row['DAY']) &
            (df['ANIMAL'] == row['ANIMAL'])
        ]
        if not linha_correspondente.empty:
            valores_droga.append(linha_correspondente['DRUG'].iloc[0])
        else:
            valores_droga.append('')

    df_final_concatenado['Droga'] = valores_droga

    # ─── Escrita no Worksheet ─────────────────────────────────
    for row in dataframe_to_rows(df_final_concatenado, index=False, header=True):
        ws.append(row)

    # ─── Formatação Visual ────────────────────────────────────

    # Merge de cabeçalhos para pares OBJ1/OBJ2
    for col_left, col_right in MERGE_COLUMN_PAIRS:
        ws.merge_cells(f'{col_left}1:{col_right}1')

    # Larguras de colunas
    for col_sep in SEPARATOR_COLUMNS:
        ws.column_dimensions[col_sep].width = SEPARATOR_COLUMN_WIDTH

    for col_left, col_right in MERGE_COLUMN_PAIRS:
        ws.column_dimensions[col_left].width = MERGE_COLUMN_WIDTH
        ws.column_dimensions[col_right].width = MERGE_COLUMN_WIDTH

    # Centralizar cabeçalho (linha 1)
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = CENTERED_ALIGNMENT

    # Centralizar dados preenchidos (linhas 2+)
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = CENTERED_ALIGNMENT

    # Renomear cabeçalhos
    for coluna_antiga, novo_nome in COLUMN_RENAMES.items():
        for cell in ws[1]:
            if cell.value == coluna_antiga:
                numero_coluna_antiga = cell.column
                break
        ws.cell(row=1, column=numero_coluna_antiga, value=novo_nome)

    # Estilizar cabeçalhos
    for cell in ws[1]:
        cell.alignment = CENTERED_ALIGNMENT
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER

    return objeto_desejado, eventos_desejados
