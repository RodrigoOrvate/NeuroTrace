"""
procurar_distvel.py — Módulo de Processamento de Distância e Velocidade

Responsável por organizar dados de distância percorrida e velocidade média
gerados pelo software Topscan (aba "Bin Measure"). Agrupa por dia em abas
separadas, detecta automaticamente todos os bins disponíveis, e calcula
a velocidade real como Distância Total / (N Bins Preenchidos * Bin Size).

Ordem das colunas de saída por aba (dia):
  Dia | Animal | [1º Xs (mm) ... Nº Xs (mm)] | Distância Total (mm) |
  Distância Total (m) | [1º Xs Vm (mm/s) ... Nº Xs Vm (mm/s)] |
  Velocidade Média (mm/s)

Regra crítica (CLAUDE.md):
  Velocidade Média = Distância Total / (Bins Preenchidos * Bin Size)
  Nunca somar ou fazer média simples de velocidades de bins diferentes.
"""

from __future__ import annotations

import logging
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# ─── Configuração de Logging ──────────────────────────────────
logger = logging.getLogger(__name__)

# ─── Constantes ───────────────────────────────────────────────
TOPSCAN_HEADER_ROW = 6
SHEET_NAME = "Bin Measure"
MEASURE_COLUMN = "Measure"
BIN_SIZE_CELL = "B5"          # célula com o Bin Size (ex: "60.0 seconds" ou 60.0)

EVENT_DISTANCE = "Mouse 1 Center Distance Traveled (apart 1.000000 second)"
EVENT_VELOCITY = "Mouse 1 Center Velocity Average (apart 1.000000 second)"

MERGE_KEYS = ["DAY", "ANIMAL"]
MM_TO_METERS_DIVISOR = 1000

# ─── Estilos do Excel ────────────────────────────────────────
CENTERED      = Alignment(horizontal="center", vertical="center")
HEADER_FONT   = Font(bold=True)
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ─── Larguras de colunas (em caracteres) ─────────────────────
WIDTH_DIA       = 8
WIDTH_ANIMAL    = 10
WIDTH_BIN_DIST  = 16
WIDTH_TOTAL_MM  = 22
WIDTH_TOTAL_M   = 18
WIDTH_BIN_VEL   = 22
WIDTH_VEL_MEDIA = 24


# ─── Funções Auxiliares ───────────────────────────────────────

def _ler_bin_size(caminho_arquivo: str) -> float:
    """Lê o valor de Bin Size diretamente da célula B5 da aba 'Bin Measure'.

    Aceita tanto valor numérico puro (60.0) quanto texto com unidade ('60.0 seconds').
    Extrai apenas a parte numérica antes do primeiro espaço e normaliza vírgula → ponto.
    Retorna 1.0 somente como último recurso de falha total.
    """
    try:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]
        valor = ws[BIN_SIZE_CELL].value
        wb.close()
        if valor is None:
            logger.warning("Célula %s está vazia — usando Bin Size = 1.0", BIN_SIZE_CELL)
            return 1.0
        # Extrai a parte numérica: converte para str, pega token antes do 1º espaço,
        # normaliza separador decimal (vírgula → ponto)
        parte_numerica = str(valor).split()[0].replace(",", ".")
        return float(parte_numerica)
    except PermissionError:
        raise  # propaga para main.py tratar com QMessageBox
    except Exception as exc:
        logger.warning(
            "Não foi possível extrair número de Bin Size na célula %s (valor=%r): %s — usando 1.0",
            BIN_SIZE_CELL, locals().get("valor"), exc,
        )
    return 1.0


def _detectar_colunas_bin(df: pd.DataFrame) -> List[str]:
    """Detecta dinamicamente todas as colunas Bin presentes no DataFrame.

    Ordena numericamente para garantir Bin1, Bin2, ..., Bin10
    (e não Bin1, Bin10, Bin2 como faria uma ordenação alfabética).
    """
    def _chave_numerica(nome: str) -> int:
        digitos = "".join(filter(str.isdigit, nome))
        return int(digitos) if digitos else 0

    colunas = [col for col in df.columns if str(col).upper().startswith("BIN")]
    return sorted(colunas, key=_chave_numerica)


def _normalize_day_column(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza a coluna DAY: remove sufixo '.0', converte para inteiro,
    e descarta linhas com valores inválidos."""
    df["DAY"] = df["DAY"].astype(str).str.replace(".0", "", regex=False)
    df["DAY"] = pd.to_numeric(df["DAY"], errors="coerce")
    return df.dropna(subset=["DAY"])


def _ordinal_pt(n: int) -> str:
    """Retorna o ordinal em português para o índice 1-based n (ex: '1º', '2º')."""
    return f"{n}º"


def _formatar_bin_size(bin_size: float) -> str:
    """Formata bin_size como inteiro se for número inteiro (ex: 60.0 → '60')."""
    return str(int(bin_size)) if bin_size == int(bin_size) else str(bin_size)


def _build_headers(n_bins: int, bin_size: float) -> Tuple[List[str], List[str]]:
    """Constrói os cabeçalhos de exibição para colunas de distância e velocidade por bin.

    Retorna (dist_headers, vel_headers), ambos com n_bins entradas.
    Exemplo para bin_size=60, n_bins=2:
        dist_headers = ['1º 60s (mm)', '2º 60s (mm)']
        vel_headers  = ['1º 60s Vm (mm/s)', '2º 60s Vm (mm/s)']
    """
    bs = _formatar_bin_size(bin_size)
    dist_headers = [f"{_ordinal_pt(i + 1)} {bs}s (mm)"       for i in range(n_bins)]
    vel_headers  = [f"{_ordinal_pt(i + 1)} {bs}s Vm (mm/s)"  for i in range(n_bins)]
    return dist_headers, vel_headers


def _processar_dia(
    df: pd.DataFrame,
    dia: int,
    colunas_bin: List[str],
    bin_size: float,
) -> pd.DataFrame:
    """Processa distância e velocidade de um dia, gerando uma linha por animal.

    Estrutura interna do DataFrame retornado (renomeada na formatação):
      DAY, ANIMAL,
      dist_Bin1 ... dist_BinN,
      Distancia_Total, Distancia_Metros,
      vel_Bin1 ... vel_BinN,
      Velocidade_Media

    Velocidade Média = Distância Total / (Bins Preenchidos * Bin Size)
    — nunca média simples das velocidades de bins individuais (CLAUDE.md).
    """
    dados_dist = df[
        (df[MEASURE_COLUMN] == EVENT_DISTANCE) & (df["DAY"] == dia)
    ].copy()

    dados_vel = df[
        (df[MEASURE_COLUMN] == EVENT_VELOCITY) & (df["DAY"] == dia)
    ].copy()

    if dados_dist.empty:
        return pd.DataFrame()

    # Converte bins para numérico
    for col in colunas_bin:
        dados_dist[col] = pd.to_numeric(dados_dist[col], errors="coerce").fillna(0)

    if not dados_vel.empty:
        for col in colunas_bin:
            dados_vel[col] = pd.to_numeric(dados_vel[col], errors="coerce").fillna(0)

    rows = []
    for _, dist_row in dados_dist.iterrows():
        animal = dist_row["ANIMAL"]

        # Distância por bin
        dist_bins = [float(dist_row[col]) for col in colunas_bin]
        dist_total = sum(dist_bins)
        bins_preenchidos = sum(1 for v in dist_bins if v > 0)

        # Velocidade por bin (evento Velocity Average do Topscan)
        vel_bins = [0.0] * len(colunas_bin)
        if not dados_vel.empty:
            vel_match = dados_vel[dados_vel["ANIMAL"] == animal]
            if not vel_match.empty:
                vel_bins = [float(vel_match.iloc[0][col]) for col in colunas_bin]

        # Velocidade média: Distância Total / (Bins Preenchidos * Bin Size)
        velocidade_media = (
            dist_total / (bins_preenchidos * bin_size)
            if bins_preenchidos > 0 and bin_size > 0
            else 0.0
        )

        row: dict = {
            "DAY":    dist_row["DAY"],
            "ANIMAL": animal,
        }
        for i, col in enumerate(colunas_bin):
            row[f"dist_{col}"] = dist_bins[i]
        row["Distancia_Total"]  = dist_total
        row["Distancia_Metros"] = dist_total / MM_TO_METERS_DIVISOR
        for i, col in enumerate(colunas_bin):
            row[f"vel_{col}"] = vel_bins[i]
        row["Velocidade_Media"] = velocidade_media

        rows.append(row)

    return pd.DataFrame(rows)


def _apply_worksheet_formatting(
    ws: Worksheet,
    colunas_bin: List[str],
    bin_size: float,
) -> None:
    """Renomeia cabeçalhos para nomes de exibição e aplica formatação visual.

    O mapa de renomeação é construído dinamicamente com base em bin_size e n_bins,
    garantindo cabeçalhos no formato '1º 60s (mm)' / '1º 60s Vm (mm/s)'.
    """
    n_bins = len(colunas_bin)
    dist_headers, vel_headers = _build_headers(n_bins, bin_size)

    # Mapa: nome interno do DataFrame → nome de exibição na planilha
    rename_map: dict = {
        "DAY":             "Dia",
        "ANIMAL":          "Animal",
        "Distancia_Total":  "Distância Total (mm)",
        "Distancia_Metros": "Distância Total (m)",
        "Velocidade_Media": "Velocidade Média (mm/s)",
    }
    for i, col in enumerate(colunas_bin):
        rename_map[f"dist_{col}"] = dist_headers[i]
        rename_map[f"vel_{col}"]  = vel_headers[i]

    # Renomeia linha de cabeçalho (row 1)
    for cell in ws[1]:
        if cell.value in rename_map:
            cell.value = rename_map[cell.value]

    # ─── Larguras dinâmicas ───────────────────────────────────
    # A=Dia, B=Animal
    ws.column_dimensions["A"].width = WIDTH_DIA
    ws.column_dimensions["B"].width = WIDTH_ANIMAL

    col_dist_inicio = 3  # coluna C
    for i in range(n_bins):
        ws.column_dimensions[get_column_letter(col_dist_inicio + i)].width = WIDTH_BIN_DIST

    col_total  = col_dist_inicio + n_bins
    col_metros = col_total + 1
    ws.column_dimensions[get_column_letter(col_total)].width  = WIDTH_TOTAL_MM
    ws.column_dimensions[get_column_letter(col_metros)].width = WIDTH_TOTAL_M

    col_vel_inicio = col_metros + 1
    for i in range(n_bins):
        ws.column_dimensions[get_column_letter(col_vel_inicio + i)].width = WIDTH_BIN_VEL

    col_vel_media = col_vel_inicio + n_bins
    ws.column_dimensions[get_column_letter(col_vel_media)].width = WIDTH_VEL_MEDIA

    # ─── Estilo de todas as células ──────────────────────────
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = CENTERED
            if cell.row == 1:
                cell.font   = HEADER_FONT
                cell.border = HEADER_BORDER


def _get_or_create_sheet(workbook: Workbook, name: str) -> Worksheet:
    """Retorna a aba existente ou cria uma nova no workbook."""
    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(title=name)


# ─── Função Principal (API Pública) ──────────────────────────

def organizar(
    caminho_arquivo: str,
    workbook: Workbook,
) -> List[int]:
    """
    Organiza dados de distância/velocidade do Topscan em abas por dia.

    Lê o Bin Size da célula B3, detecta automaticamente todas as colunas
    Bin, e calcula a velocidade real (não uma média de médias).

    Cada aba representa um dia e contém:
      Dia | Animal | Dist por Bin... | Distância Total (mm) |
      Distância Total (m) | Vel por Bin... | Velocidade Média (mm/s)

    Args:
        caminho_arquivo: Caminho completo do arquivo Excel do Topscan.
        workbook:        Workbook do openpyxl para escrita.

    Returns:
        Lista com os dias únicos processados (inteiros), ou lista vazia em caso de erro.
    """
    bin_size = _ler_bin_size(caminho_arquivo)
    logger.info("Bin Size lido da planilha: %.4f s", bin_size)

    try:
        df = pd.read_excel(
            caminho_arquivo,
            header=TOPSCAN_HEADER_ROW,
            sheet_name=SHEET_NAME,
        )
    except PermissionError:
        raise  # main.py trata com QMessageBox
    except Exception as exc:
        logger.error("Erro ao ler a aba '%s': %s", SHEET_NAME, exc)
        return []

    colunas_bin = _detectar_colunas_bin(df)
    if not colunas_bin:
        logger.error("Nenhuma coluna Bin encontrada na aba '%s'.", SHEET_NAME)
        return []
    logger.info("Bins detectados (%d): %s", len(colunas_bin), colunas_bin)

    df = _normalize_day_column(df)
    unique_days = sorted(df["DAY"].unique().astype(int))

    for dia in unique_days:
        ws = _get_or_create_sheet(workbook, str(dia))

        df_dia = _processar_dia(df, dia, colunas_bin, bin_size)

        if df_dia.empty:
            logger.debug("Dia %d sem dados de distância, pulando.", dia)
            continue

        for row in dataframe_to_rows(df_dia, index=False, header=True):
            ws.append(row)

        _apply_worksheet_formatting(ws, colunas_bin, bin_size)
        logger.info(
            "Dia %d processado com sucesso (%d bins, bin_size=%.4f s).",
            dia, len(colunas_bin), bin_size,
        )

    return unique_days
