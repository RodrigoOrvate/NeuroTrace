import sys
import os
import platform
import subprocess
from functools import partial
import pandas as pd
import openpyxl
from qt_compat import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QScrollArea, QFrame, QFileDialog,
    QMessageBox, QGraphicsDropShadowEffect, QSpinBox, QSizePolicy,
    QGroupBox, QGridLayout, QSpacerItem,
    Qt, QPropertyAnimation, QEasingCurve, QSize, pyqtSignal, QTimer,
    QFont, QIcon, QColor, QPalette, QFontDatabase, QPixmap,
)

# Importando as funções de processamento (inalteradas)
from procurar_objeto import procurar
from procurar_distvel import organizar
from updater import (
    check_for_updates, check_whats_new, check_internet,
    CURRENT_VERSION, GITHUB_REPO_URL, _open_url,
)

def resource_path(relative_path: str) -> str:
    """Resolve caminho de recurso para PyInstaller bundle ou dev."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)

# ─── Paleta de Cores ───────────────────────────────────────────
COLORS = {
    "bg":           "#0f0f1a",
    "card":         "#1a1a2e",
    "card_border":  "#2d2d4a",
    "accent":       "#ab3d4c",
    "accent_hover": "#c9505f",
    "accent_dark":  "#8a2e3b",
    "danger":       "#ff4757",
    "danger_hover": "#ff6b7a",
    "text":         "#e8e8f0",
    "text_muted":   "#8888aa",
    "input_bg":     "#12122a",
    "input_border": "#3a3a5c",
    "input_focus":  "#ab3d4c",
    "surface":      "#16162e",
    "success":      "#ab3d4c",
    "warning":      "#ffa502",
    "scroll_bg":    "#1a1a2e",
    "scroll_handle":"#3a3a5c",
}

# ─── Stylesheet Global ────────────────────────────────────────
GLOBAL_STYLESHEET = f"""
QMainWindow {{
    background-color: {COLORS['bg']};
}}
QWidget {{
    color: {COLORS['text']};
    font-family: 'Segoe UI', 'Roboto', 'Arial', sans-serif;
}}
QLabel {{
    color: {COLORS['text']};
    background: transparent;
}}
QLineEdit {{
    background-color: {COLORS['input_bg']};
    border: 1.5px solid {COLORS['input_border']};
    border-radius: 8px;
    padding: 8px 14px;
    color: {COLORS['text']};
    font-size: 13px;
    selection-background-color: {COLORS['accent']};
}}
QLineEdit:focus {{
    border-color: {COLORS['input_focus']};
    background-color: #14142e;
}}
QLineEdit:read-only {{
    background-color: {COLORS['surface']};
    border-color: {COLORS['card_border']};
    color: {COLORS['text_muted']};
}}
QSpinBox {{
    background-color: {COLORS['input_bg']};
    border: 1.5px solid {COLORS['input_border']};
    border-radius: 8px;
    padding: 8px 14px;
    color: {COLORS['text']};
    font-size: 13px;
    min-width: 80px;
}}
QSpinBox:focus {{
    border-color: {COLORS['input_focus']};
}}
QSpinBox::up-button, QSpinBox::down-button {{
    background-color: {COLORS['card']};
    border: none;
    border-radius: 4px;
    width: 20px;
}}
QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
    background-color: {COLORS['accent']};
}}
QPushButton {{
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 600;
    letter-spacing: 0.3px;
}}
QScrollArea {{
    border: none;
    background: transparent;
}}
QScrollBar:vertical {{
    background: {COLORS['scroll_bg']};
    width: 8px;
    border-radius: 4px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {COLORS['scroll_handle']};
    min-height: 30px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical:hover {{
    background: {COLORS['accent']};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0;
}}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
    background: none;
}}
QGroupBox {{
    background-color: {COLORS['card']};
    border: 1px solid {COLORS['card_border']};
    border-radius: 12px;
    margin-top: 12px;
    padding: 14px;
    padding-top: 14px;
    font-size: 14px;
    font-weight: 600;
    color: {COLORS['accent']};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 6px 16px;
    background-color: {COLORS['card']};
    border: 1px solid {COLORS['card_border']};
    border-radius: 8px;
    color: {COLORS['accent']};
    left: 14px;
}}
"""

def make_accent_button(text, icon_text=""):
    """Cria um botão com estilo accent (vermelho laboratorial)."""
    btn = QPushButton(f" {icon_text} {text} " if icon_text else f" {text} ")
    btn.setStyleSheet(f"""
        QPushButton {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {COLORS['accent']}, stop:1 {COLORS['accent_dark']});
            color: #ffffff;
            font-weight: 700;
            font-size: 13px;
            padding: 10px 24px;
            border-radius: 8px;
        }}
        QPushButton:hover {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {COLORS['accent_hover']}, stop:1 {COLORS['accent']});
        }}
        QPushButton:pressed {{
            background: {COLORS['accent_dark']};
        }}
        QPushButton:disabled {{
            background: {COLORS['card_border']};
            color: {COLORS['text_muted']};
        }}
    """)
    btn.setCursor(Qt.PointingHandCursor)
    btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
    shadow = QGraphicsDropShadowEffect()
    shadow.setBlurRadius(18)
    shadow.setOffset(0, 4)
    shadow.setColor(QColor(171, 61, 76, 60))
    btn.setGraphicsEffect(shadow)
    return btn

def make_danger_button(text, icon_text=""):
    """Cria um botão com estilo danger (vermelho)."""
    btn = QPushButton(f" {icon_text} {text} " if icon_text else f" {text} ")
    btn.setStyleSheet(f"""
        QPushButton {{
            background: transparent;
            color: {COLORS['danger']};
            border: 1.5px solid {COLORS['danger']};
            font-weight: 600;
            font-size: 12px;
            padding: 6px 14px;
            border-radius: 6px;
        }}
        QPushButton:hover {{
            background: {COLORS['danger']};
            color: white;
        }}
    """)
    btn.setCursor(Qt.PointingHandCursor)
    btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
    return btn

def make_secondary_button(text):
    """Cria um botão secundário (outline)."""
    btn = QPushButton(f" {text} ")
    btn.setStyleSheet(f"""
        QPushButton {{
            background: transparent;
            color: {COLORS['accent']};
            border: 1.5px solid {COLORS['accent']};
            font-weight: 600;
            font-size: 13px;
            padding: 10px 22px;
            border-radius: 8px;
        }}
        QPushButton:hover {{
            background: {COLORS['accent']};
            color: #0a0a1a;
        }}
        QPushButton:disabled {{
            border-color: {COLORS['card_border']};
            color: {COLORS['text_muted']};
        }}
    """)
    btn.setCursor(Qt.PointingHandCursor)
    btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
    return btn

def make_github_button():
    """Cria o botão do GitHub com ícone SVG embutido via pixmap."""
    btn = QPushButton()
    btn.setFixedSize(38, 38)
    btn.setToolTip("Abrir repositório no GitHub")
    btn.setCursor(Qt.PointingHandCursor)
    btn.setStyleSheet(f"""
        QPushButton {{
            background: transparent;
            border: 1.5px solid {COLORS['card_border']};
            border-radius: 8px;
            padding: 0px;
            font-size: 18px;
        }}
        QPushButton:hover {{
            border-color: {COLORS['text_muted']};
            background: {COLORS['card']};
        }}
        QPushButton:disabled {{
            border-color: {COLORS['card_border']};
            opacity: 0.4;
        }}
    """)
    # Usa o logo do GitHub em unicode como fallback elegante
    btn.setText("⌥")  # será substituído abaixo

    # Tenta desenhar ícone SVG do GitHub via QPixmap/QPainter
    try:
        from qt_compat import QPixmap, QPainter
        # SVG simplificado do logo GitHub (Octocat simplificado como ícone)
        github_svg = b"""<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="20" height="20">
  <path fill="#8888aa" d="M12 0C5.37 0 0 5.37 0 12c0 5.31 3.435 9.795 8.205 11.385
    .6.105.825-.255.825-.57 0-.285-.015-1.23-.015-2.235-3.015.555-3.795-.735-4.035-1.41
    -.135-.345-.72-1.41-1.23-1.695-.42-.225-1.02-.78-.015-.795.945-.015 1.62.87 1.845 1.23
    1.08 1.815 2.805 1.305 3.495.99.105-.78.42-1.305.765-1.605-2.67-.3-5.46-1.335-5.46-5.925
    0-1.305.465-2.385 1.23-3.225-.12-.3-.54-1.53.12-3.18 0 0 1.005-.315 3.3 1.23
    .96-.27 1.98-.405 3-.405s2.04.135 3 .405c2.295-1.56 3.3-1.23 3.3-1.23
    .66 1.65.24 2.88.12 3.18.765.84 1.23 1.905 1.23 3.225 0 4.605-2.805 5.625-5.475 5.925
    .435.375.81 1.095.81 2.22 0 1.605-.015 2.895-.015 3.3 0 .315.225.69.825.57
    A12.02 12.02 0 0 0 24 12c0-6.63-5.37-12-12-12z"/>
</svg>"""
        import tempfile as _tf, os as _os
        tmp = _tf.NamedTemporaryFile(suffix=".svg", delete=False)
        tmp.write(github_svg)
        tmp.close()
        pix = QPixmap(tmp.name)
        _os.unlink(tmp.name)
        if not pix.isNull():
            btn.setIcon(QIcon(pix))
            btn.setIconSize(QSize(20, 20))
            btn.setText("")
        else:
            btn.setText("🐙")  # Octocat Unicode fallback
    except Exception:
        btn.setText("🐙")  # Octocat Unicode fallback

    return btn


class PlaceholderLineEdit(QLineEdit):
    """QLineEdit personalizado com placeholder que desaparece ao focar."""
    def __init__(self, placeholder="", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)


class ConjuntoCard(QFrame):
    """Card visual para cada conjunto de planilha (par de objetos)."""
    def __init__(self, numero, parent=None):
        super().__init__(parent)
        self.numero = numero
        self._setup_ui()
        self._setup_style()

    def _setup_style(self):
        self.setStyleSheet(f"""
            ConjuntoCard {{
                background-color: {COLORS['surface']};
                border: 1px solid {COLORS['card_border']};
                border-radius: 10px;
            }}
        """)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 14)
        layout.setSpacing(10)

        title = QLabel(f"📋 Planilha {self.numero}")
        title.setStyleSheet(f"""
            font-size: 14px; font-weight: 700;
            color: {COLORS['accent']}; padding-bottom: 4px;
        """)
        layout.addWidget(title)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet(f"background-color: {COLORS['card_border']}; max-height: 1px;")
        layout.addWidget(separator)

        grid = QGridLayout()
        grid.setSpacing(8)

        lbl_par1 = QLabel("Par Objeto 1:")
        lbl_par1.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px;")
        self.objeto1_entry = PlaceholderLineEdit("Ex: A")
        grid.addWidget(lbl_par1, 0, 0)
        grid.addWidget(self.objeto1_entry, 0, 1)

        lbl_par2 = QLabel("Par Objeto 2:")
        lbl_par2.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px;")
        self.objeto2_entry = PlaceholderLineEdit("Ex: B")
        grid.addWidget(lbl_par2, 0, 2)
        grid.addWidget(self.objeto2_entry, 0, 3)

        lbl_obj1 = QLabel("OBJ 1:")
        lbl_obj1.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px;")
        self.obj1_entry = PlaceholderLineEdit("Ex: 1")
        grid.addWidget(lbl_obj1, 1, 0)
        grid.addWidget(self.obj1_entry, 1, 1)

        lbl_obj2 = QLabel("OBJ 2:")
        lbl_obj2.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 12px;")
        self.obj2_entry = PlaceholderLineEdit("Ex: 2")
        grid.addWidget(lbl_obj2, 1, 2)
        grid.addWidget(self.obj2_entry, 1, 3)

        layout.addLayout(grid)

    def get_values(self):
        return (
            self.objeto1_entry.text().strip(),
            self.objeto2_entry.text().strip(),
            self.obj1_entry.text().strip(),
            self.obj2_entry.text().strip(),
        )

    def all_filled(self):
        return all(v != "" for v in self.get_values())


class MainWindow(QMainWindow):
    """Janela principal do aplicativo NeuroTrace."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("NeuroTrace — Topscan Data Organizer")
        self.setMinimumSize(820, 780)
        self.resize(860, 800)

        # Estado
        self.caminho_arquivo1 = ""
        self.caminho_arquivo2 = ""
        self.conjuntos_cards  = []
        self.global_workbook  = openpyxl.Workbook()
        self.global_excel_filename_obj    = "dados_filtrados_obj.xlsx"
        self.global_excel_filename_distvel = "dados_filtrados_distvel.xlsx"
        self.colunas_desejadas = [
            'DAY', 'ANIMAL', 'OBJECTS', 'Total Bouts',
            'Total Duration(Second)', 'Latency(Second)',
            'Ending time(Second) of First Bout'
        ]
        self.pares_objetos = set()
        self.objs          = set()

        # Estado de internet (atualizado periodicamente)
        self._internet_ok = False

        # Ícone
        try:
            caminho_icone = resource_path("memorylab.ico")
            if os.path.exists(caminho_icone):
                self.setWindowIcon(QIcon(caminho_icone))
        except Exception:
            pass

        self._build_ui()

        # Timer para validação contínua de botões
        self.validation_timer = QTimer(self)
        self.validation_timer.timeout.connect(self._validate_buttons)
        self.validation_timer.start(300)

        # Timer para checar internet a cada 10 segundos
        self._internet_timer = QTimer(self)
        self._internet_timer.timeout.connect(self._check_internet_status)
        self._internet_timer.start(10_000)

        # Centraliza na tela
        self._center_on_screen()

        # Na inicialização: verifica internet, depois dispara rotinas online
        QTimer.singleShot(500,  self._check_internet_status)
        QTimer.singleShot(2000, self._on_startup_online_checks)

    def _center_on_screen(self):
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width()  - self.width())  // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)

    # ─── Verificação de Internet ───────────────────────────────

    def _check_internet_status(self):
        """Checa a conectividade em thread separada para não travar a UI."""
        from qt_compat import QThread

        class _Checker(QThread):
            result = pyqtSignal(bool)
            def run(self):
                self.result.emit(check_internet())

        checker = _Checker()
        checker.result.connect(self._apply_internet_status)
        checker.start()
        # Mantém referência para não ser coletado pelo GC
        self._inet_checker = checker

    def _apply_internet_status(self, online: bool):
        """Atualiza o estado de internet e habilita/desabilita botões dependentes."""
        self._internet_ok = online
        self.github_btn.setEnabled(online)
        self.update_btn.setEnabled(online)

        # Tooltip explicativo quando offline
        if online:
            self.github_btn.setToolTip("Abrir repositório no GitHub")
            self.update_btn.setToolTip("")
        else:
            self.github_btn.setToolTip("Sem conexão com a internet")
            self.update_btn.setToolTip("Sem conexão com a internet")

    def _on_startup_online_checks(self):
        """Disparado após a janela carregar: verifica atualização e novidades."""
        if self._internet_ok:
            # "O que há de novo" — só aparece se for primeira abertura da versão
            check_whats_new(self)
            # Verifica se há versão ainda mais nova disponível
            QTimer.singleShot(500, lambda: check_for_updates(self, silent=True))

    # ─── Build da UI ───────────────────────────────────────────

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(28, 20, 28, 20)
        main_layout.setSpacing(16)

        # ─── Header ───
        header = QLabel("NeuroTrace")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet(f"""
            font-size: 26px; font-weight: 800;
            letter-spacing: 3px; color: {COLORS['accent']}; padding: 6px 0;
        """)
        main_layout.addWidget(header)

        subtitle = QLabel(f"Organizador de dados Topscan · v{CURRENT_VERSION}")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet(f"""
            font-size: 12px; color: {COLORS['text_muted']};
            letter-spacing: 1px; margin-bottom: 8px;
        """)
        main_layout.addWidget(subtitle)

        # ─── Seção 1: Arquivos ───
        files_group = QGroupBox("Arquivos de Entrada")
        files_layout = QVBoxLayout(files_group)
        files_layout.setContentsMargins(14, 30, 14, 14)
        files_layout.setSpacing(10)

        row1 = QHBoxLayout()
        row1.setSpacing(8)
        self.pesquisar1_btn = make_accent_button("Pesquisar Arquivo (OBJ)", "📂")
        self.pesquisar1_btn.clicked.connect(self._pesquisar_arquivo1)
        row1.addWidget(self.pesquisar1_btn)
        self.caminho_entry1 = QLineEdit()
        self.caminho_entry1.setReadOnly(True)
        self.caminho_entry1.setPlaceholderText("Nenhum arquivo selecionado...")
        row1.addWidget(self.caminho_entry1, stretch=1)
        self.limpar1_btn = make_danger_button("✕")
        self.limpar1_btn.setFixedWidth(36)
        self.limpar1_btn.clicked.connect(self._limpar_entry1)
        row1.addWidget(self.limpar1_btn)
        files_layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.setSpacing(8)
        self.pesquisar2_btn = make_accent_button("Pesquisar Arquivo (DIST/VEL)", "📂")
        self.pesquisar2_btn.clicked.connect(self._pesquisar_arquivo2)
        row2.addWidget(self.pesquisar2_btn)
        self.caminho_entry2 = QLineEdit()
        self.caminho_entry2.setReadOnly(True)
        self.caminho_entry2.setPlaceholderText("Nenhum arquivo selecionado...")
        row2.addWidget(self.caminho_entry2, stretch=1)
        self.limpar2_btn = make_danger_button("✕")
        self.limpar2_btn.setFixedWidth(36)
        self.limpar2_btn.clicked.connect(self._limpar_entry2)
        row2.addWidget(self.limpar2_btn)
        files_layout.addLayout(row2)

        files_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        main_layout.addWidget(files_group)

        # ─── Seção 2: Configuração de Conjuntos ───
        config_group = QGroupBox("Configuração dos Conjuntos")
        config_layout = QVBoxLayout(config_group)
        config_layout.setContentsMargins(14, 30, 14, 14)
        config_layout.setSpacing(12)

        config_row1 = QHBoxLayout()
        config_row1.setSpacing(12)

        lbl_qty = QLabel("Quantidade de planilhas:")
        lbl_qty.setStyleSheet(f"font-weight: 600; font-size: 13px; color: {COLORS['text']};")
        config_row1.addWidget(lbl_qty)

        self.quantidade_spin = QSpinBox()
        self.quantidade_spin.setRange(0, 100)
        self.quantidade_spin.setValue(0)
        self.quantidade_spin.setFixedWidth(90)
        self.quantidade_spin.setFixedHeight(36)
        config_row1.addWidget(self.quantidade_spin)

        config_row1.addSpacing(10)

        self.criar_btn = make_accent_button("Criar Conjuntos", "➕")
        self.criar_btn.setEnabled(False)
        self.criar_btn.clicked.connect(self._criar_conjuntos)
        config_row1.addWidget(self.criar_btn)

        config_row1.addSpacing(6)

        self.ver_objetos_btn = make_secondary_button("Ver Objetos 🔍")
        self.ver_objetos_btn.clicked.connect(self._atualizar_rotulos)
        config_row1.addWidget(self.ver_objetos_btn)

        config_row1.addStretch(1)
        config_layout.addLayout(config_row1)

        self.pares_label = QLabel("")
        self.pares_label.setWordWrap(True)
        self.pares_label.setStyleSheet(f"font-size: 12px; color: {COLORS['text_muted']}; padding: 2px 0;")
        config_layout.addWidget(self.pares_label)

        self.objs_label = QLabel("")
        self.objs_label.setWordWrap(True)
        self.objs_label.setStyleSheet(f"font-size: 12px; color: {COLORS['text_muted']}; padding: 2px 0;")
        config_layout.addWidget(self.objs_label)

        self.limite_label = QLabel("")
        self.limite_label.setStyleSheet(f"color: {COLORS['danger']}; font-size: 12px;")
        config_layout.addWidget(self.limite_label)

        config_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        main_layout.addWidget(config_group)

        # ─── Seção 3: Scroll de Conjuntos ───
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setMinimumHeight(100)
        self.scroll_area.setStyleSheet(f"""
            QScrollArea {{
                background: {COLORS['bg']};
                border: 1px solid {COLORS['card_border']};
                border-radius: 10px;
            }}
        """)
        self.scroll_widget = QWidget()
        self.scroll_widget.setStyleSheet(f"background: {COLORS['bg']};")
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_layout.setContentsMargins(8, 8, 8, 8)
        self.scroll_layout.setSpacing(10)
        self.scroll_layout.addStretch()
        self.scroll_area.setWidget(self.scroll_widget)
        main_layout.addWidget(self.scroll_area, stretch=1)

        # ─── Seção 4: Botões de Ação ───
        action_frame = QFrame()
        action_frame.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['card']};
                border: 1px solid {COLORS['card_border']};
                border-radius: 12px;
                padding: 8px;
            }}
        """)
        action_layout = QHBoxLayout(action_frame)
        action_layout.setContentsMargins(16, 10, 16, 10)
        action_layout.setSpacing(12)

        self.procurar_obj_btn = make_accent_button("Procurar Objetos", "🔬")
        self.procurar_obj_btn.setEnabled(False)
        self.procurar_obj_btn.clicked.connect(self._procurar_objetos)
        action_layout.addWidget(self.procurar_obj_btn)

        self.organizar_distvel_btn = make_accent_button("Organizar Dist/Vel", "📊")
        self.organizar_distvel_btn.setEnabled(False)
        self.organizar_distvel_btn.clicked.connect(self._organizar_distvel)
        action_layout.addWidget(self.organizar_distvel_btn)

        self.reiniciar_btn = make_secondary_button("Reiniciar 🔄")
        self.reiniciar_btn.clicked.connect(self._reiniciar_programa)
        action_layout.addWidget(self.reiniciar_btn)

        # Botão Atualizar — dependente de internet
        self.update_btn = make_secondary_button("Atualizar 🔄")
        self.update_btn.setEnabled(False)  # começa desabilitado até confirmar internet
        self.update_btn.clicked.connect(self._check_updates_manual)
        action_layout.addWidget(self.update_btn)

        action_layout.addStretch()

        # Botão GitHub — dependente de internet
        self.github_btn = make_github_button()
        self.github_btn.setEnabled(False)  # começa desabilitado até confirmar internet
        self.github_btn.clicked.connect(self._open_github)
        action_layout.addWidget(self.github_btn)

        main_layout.addWidget(action_frame)

    # ─── Ações de Arquivo ──────────────────────────────────────

    def _pesquisar_arquivo1(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Arquivo (OBJ)", "",
            "Arquivos Excel (*.xlsx *.xls);;Todos (*.*)"
        )
        if filename:
            self.caminho_arquivo1 = filename
            self.caminho_entry1.setText(filename)
            self._atualizar_rotulos()

    def _pesquisar_arquivo2(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Arquivo (DIST/VEL)", "",
            "Arquivos Excel (*.xlsx *.xls);;Todos (*.*)"
        )
        if filename:
            self.caminho_arquivo2 = filename
            self.caminho_entry2.setText(filename)

    def _limpar_entry1(self):
        self.caminho_arquivo1 = ""
        self.caminho_entry1.clear()
        self.quantidade_spin.setValue(0)
        self.pares_label.setText("")
        self.objs_label.setText("")
        self._clear_conjuntos()

    def _limpar_entry2(self):
        self.caminho_arquivo2 = ""
        self.caminho_entry2.clear()

    # ─── Gerenciamento de Conjuntos ────────────────────────────

    def _criar_conjuntos(self):
        num = self.quantidade_spin.value()
        if num == 0:
            self.limite_label.setText("⚠ Mínimo de planilhas é 1.")
            return
        if num > 100:
            self.limite_label.setText("⚠ Limite máximo de 100 planilhas.")
            return
        self.limite_label.setText("")
        self._clear_conjuntos()
        self.global_workbook = openpyxl.Workbook()
        for i in range(1, num + 1):
            card = ConjuntoCard(i)
            self.scroll_layout.insertWidget(self.scroll_layout.count() - 1, card)
            self.conjuntos_cards.append(card)

    def _clear_conjuntos(self):
        for card in self.conjuntos_cards:
            card.setParent(None)
            card.deleteLater()
        self.conjuntos_cards.clear()

    # ─── Validação de Botões ───────────────────────────────────

    def _validate_buttons(self):
        self.criar_btn.setEnabled(bool(self.caminho_arquivo1))

        can_search_obj = (
            bool(self.caminho_arquivo1)
            and len(self.conjuntos_cards) > 0
            and all(card.all_filled() for card in self.conjuntos_cards)
        )
        self.procurar_obj_btn.setEnabled(can_search_obj)
        self.organizar_distvel_btn.setEnabled(bool(self.caminho_arquivo2))

        # Botões dependentes de internet são controlados por _apply_internet_status

    # ─── Processamento ─────────────────────────────────────────

    def _procurar_objetos(self):
        if 'Sheet' in self.global_workbook.sheetnames:
            del self.global_workbook['Sheet']
        try:
            for card in self.conjuntos_cards:
                obj1, obj2, o1, o2 = card.get_values()
                procurar(
                    obj1.upper(), obj2.upper(), o1.upper(), o2.upper(),
                    self.caminho_arquivo1, self.global_workbook, self.colunas_desejadas
                )
            self.global_workbook.save(self.global_excel_filename_obj)
            self._mostrar_sucesso()
        except Exception as e:
            self._mostrar_erro(f"Erro ao processar objetos:\n{str(e)}")

    def _organizar_distvel(self):
        self.global_workbook = openpyxl.Workbook()
        try:
            organizar(self.caminho_arquivo2, self.global_workbook)
            if 'Sheet' in self.global_workbook.sheetnames:
                del self.global_workbook['Sheet']
            self.global_workbook.save(self.global_excel_filename_distvel)
            self._mostrar_sucesso()
        except Exception as e:
            self._mostrar_erro(f"Erro ao organizar Dist/Vel:\n{str(e)}")

    # ─── Leitura de Objetos ────────────────────────────────────

    def _procurar_colunas(self, caminho):
        try:
            df      = pd.read_excel(caminho, header=6)
            objetos = set(df['OBJECTS'].astype(str))
            events  = set(df['Events'].astype(str))

            self.pares_objetos.clear()
            for obj in objetos:
                if obj.strip() and obj != 'nan':
                    pares = obj.split(' & ')
                    self.pares_objetos.update(pares)

            self.objs.clear()
            for event in events:
                if "OBJ" in event:
                    try:
                        o = event.split("OBJ")[1].split()[0]
                        self.objs.add(o)
                    except IndexError:
                        pass
        except Exception as e:
            self.pares_label.setText(f"⚠ Erro ao ler arquivo: {e}")
            self.objs_label.setText("")

    def _atualizar_rotulos(self):
        if self.caminho_arquivo1:
            self._procurar_colunas(self.caminho_arquivo1)
            pares_text = ', '.join(sorted(self.pares_objetos))
            objs_text  = ', '.join(sorted(self.objs))
            self.pares_label.setText(f"🔹 Pares de Objetos: {pares_text}" if pares_text else "Nenhum par encontrado.")
            self.objs_label.setText(f"🔹 OBJs: {objs_text}" if objs_text else "Nenhum OBJ encontrado.")
        else:
            self.pares_label.setText("⚠ Selecione um arquivo OBJ primeiro.")
            self.objs_label.setText("")

    # ─── Diálogos ──────────────────────────────────────────────

    def _mostrar_sucesso(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Sucesso")
        msg.setText("✅ Dados filtrados com sucesso!\nVerifique o arquivo gerado.")
        msg.setStyleSheet(f"""
            QMessageBox {{ background-color: {COLORS['card']}; }}
            QMessageBox QLabel {{ color: {COLORS['text']}; font-size: 14px; padding: 12px; }}
            QPushButton {{
                background: {COLORS['accent']}; color: #0a0a1a;
                font-weight: 700; padding: 8px 28px;
                border-radius: 6px; font-size: 13px;
            }}
            QPushButton:hover {{ background: {COLORS['accent_hover']}; }}
        """)
        msg.exec_()
        self._abrir_arquivo_excel()

    def _mostrar_erro(self, mensagem):
        msg = QMessageBox(self)
        msg.setWindowTitle("Erro")
        msg.setIcon(QMessageBox.Warning)
        msg.setText(mensagem)
        msg.setStyleSheet(f"""
            QMessageBox {{ background-color: {COLORS['card']}; }}
            QMessageBox QLabel {{ color: {COLORS['text']}; font-size: 13px; padding: 10px; }}
            QPushButton {{
                background: {COLORS['danger']}; color: white;
                font-weight: 700; padding: 8px 28px;
                border-radius: 6px; font-size: 13px;
            }}
        """)
        msg.exec_()

    def _abrir_arquivo_excel(self):
        f_obj = self.global_excel_filename_obj
        f_dv  = self.global_excel_filename_distvel
        arquivo = None
        if os.path.exists(f_obj) and os.path.exists(f_dv):
            arquivo = f_obj if os.path.getmtime(f_obj) > os.path.getmtime(f_dv) else f_dv
        elif os.path.exists(f_obj):
            arquivo = f_obj
        elif os.path.exists(f_dv):
            arquivo = f_dv
        if arquivo:
            self._open_file_crossplatform(arquivo)

    @staticmethod
    def _open_file_crossplatform(filepath: str):
        try:
            if sys.platform == 'win32':
                os.startfile(filepath)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', filepath])
            else:
                subprocess.Popen(['xdg-open', filepath])
        except Exception:
            pass

    def _reiniciar_programa(self):
        try:
            if getattr(sys, 'frozen', False):
                subprocess.Popen([sys.executable])
            else:
                python = sys.executable
                os.chdir(os.path.dirname(os.path.abspath(sys.argv[0])))
                subprocess.Popen([python] + sys.argv)
            QApplication.quit()
        except Exception as e:
            self._mostrar_erro(f"Erro ao reiniciar: {e}")

    def _check_updates_manual(self):
        """Verifica atualizações manualmente (mostra resultado ao usuário)."""
        check_for_updates(self, silent=False)

    def _open_github(self):
        """Abre o repositório do GitHub no navegador padrão."""
        _open_url(GITHUB_REPO_URL)


# ─── Ponto de Entrada ─────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication.instance()
    if not app:
        app = QApplication(sys.argv)
    
    app.setStyle("Fusion")

    palette = QPalette()
    palette.setColor(QPalette.Window,          QColor(COLORS['bg']))
    palette.setColor(QPalette.WindowText,      QColor(COLORS['text']))
    palette.setColor(QPalette.Base,            QColor(COLORS['input_bg']))
    palette.setColor(QPalette.AlternateBase,   QColor(COLORS['card']))
    palette.setColor(QPalette.ToolTipBase,     QColor(COLORS['card']))
    palette.setColor(QPalette.ToolTipText,     QColor(COLORS['text']))
    palette.setColor(QPalette.Text,            QColor(COLORS['text']))
    palette.setColor(QPalette.Button,          QColor(COLORS['card']))
    palette.setColor(QPalette.ButtonText,      QColor(COLORS['text']))
    palette.setColor(QPalette.Highlight,       QColor(COLORS['accent']))
    palette.setColor(QPalette.HighlightedText, QColor("#0a0a1a"))
    app.setPalette(palette)
    app.setStyleSheet(GLOBAL_STYLESHEET)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
